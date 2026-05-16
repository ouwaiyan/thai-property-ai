import csv
import os
import re
import uuid as _uuid
from uuid import UUID

import openpyxl
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.import_error import ImportError
from app.models.import_job import ImportJob
from app.models.property import Property
from app.models.user import User
from app.utils.geo import validate_latlng

COLUMN_DETECT_MAP: dict[str, str] = {
    # Chinese column names
    "房源编号": "property_code",
    "编号": "property_code",
    "名称": "name",
    "房源名称": "name",
    "楼盘名称": "building_name",
    "楼盘": "building_name",
    "地址": "address",
    "详细地址": "address",
    "纬度": "latitude",
    "经度": "longitude",
    "区": "district",
    "区域": "district",
    "片区": "area",
    "最近bts": "nearest_bts",
    "bts": "nearest_bts",
    "最近mrt": "nearest_mrt",
    "mrt": "nearest_mrt",
    "卧室数": "bedroom_count",
    "卧室": "bedroom_count",
    "卫生间数": "bathroom_count",
    "卫生间": "bathroom_count",
    "面积": "size_sqm",
    "面积(平方米)": "size_sqm",
    "月租": "monthly_rent",
    "月租(泰铢)": "monthly_rent",
    "租金": "monthly_rent",
    "押金月数": "deposit_months",
    "押金(月)": "deposit_months",
    "状态": "status",
    "房源状态": "status",
    "可入住日期": "available_date",
    "入住日期": "available_date",
    "允许宠物": "pet_allowed",
    "宠物": "pet_allowed",
    "联系人": "contact_person",
    "联系": "contact_person",
    "联系电话": "contact_phone",
    "电话": "contact_phone",
    "line": "contact_line",
    "line id": "contact_line",
    "描述": "description",
    "备注": "internal_note",
    "内部备注": "internal_note",
    "标签": "tags",
    # English column names
    "property_code": "property_code",
    "code": "property_code",
    "name": "name",
    "building_name": "building_name",
    "building": "building_name",
    "address": "address",
    "latitude": "latitude",
    "lat": "latitude",
    "longitude": "longitude",
    "lng": "longitude",
    "lon": "longitude",
    "district": "district",
    "area": "area",
    "nearest_bts": "nearest_bts",
    "bts": "nearest_bts",
    "nearest_mrt": "nearest_mrt",
    "mrt": "nearest_mrt",
    "bedroom_count": "bedroom_count",
    "bedroom": "bedroom_count",
    "bedrooms": "bedroom_count",
    "bathroom_count": "bathroom_count",
    "bathroom": "bathroom_count",
    "bathrooms": "bathroom_count",
    "size_sqm": "size_sqm",
    "size": "size_sqm",
    "monthly_rent": "monthly_rent",
    "rent": "monthly_rent",
    "price": "monthly_rent",
    "deposit_months": "deposit_months",
    "deposit": "deposit_months",
    "status": "status",
    "available_date": "available_date",
    "pet_allowed": "pet_allowed",
    "pets": "pet_allowed",
    "contact_person": "contact_person",
    "contact": "contact_person",
    "contact_phone": "contact_phone",
    "phone": "contact_phone",
    "contact_line": "contact_line",
    "description": "description",
    "internal_note": "internal_note",
    "note": "internal_note",
    "tags": "tags",
}

REQUIRED_FIELDS = {"name", "address", "monthly_rent", "status"}
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_FILE_SIZE = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024

# Fields that map to Property model columns (for convert-to-Property step)
PROPERTY_FIELDS = {
    "property_code", "name", "building_name", "address", "latitude", "longitude",
    "district", "area", "nearest_bts", "nearest_mrt", "bedroom_count", "bathroom_count",
    "size_sqm", "monthly_rent", "deposit_months", "status", "available_date",
    "pet_allowed", "contact_person", "contact_phone", "contact_line",
    "description", "internal_note", "tags",
}


def _clean_price(value: str | None) -> int | None:
    if value is None:
        return None
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return int(float(cleaned))
    except (ValueError, TypeError):
        return None


def _clean_float(value: str | None) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _clean_int(value: str | None) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    cleaned = re.sub(r"[^\d\-]", "", str(value))
    try:
        return int(cleaned)
    except (ValueError, TypeError):
        return None


def _clean_bool(value: str | None) -> bool:
    if value is None or str(value).strip() == "":
        return False
    v = str(value).strip().lower()
    return v in ("yes", "true", "1", "是", "y", "有", "允许")


def _detect_column(field_name: str) -> str | None:
    key = field_name.strip().lower()
    return COLUMN_DETECT_MAP.get(key, COLUMN_DETECT_MAP.get(field_name.strip()))


def _resolve_path(file_path: str) -> str:
    """Convert a relative file path stored on ImportJob to an absolute filesystem path."""
    return os.path.abspath(os.path.join(settings.UPLOAD_DIR, "..", "imports", os.path.basename(file_path)))


def _read_headers(file_path: str) -> list[str]:
    abs_path = _resolve_path(file_path)
    ext = os.path.splitext(abs_path)[1].lower()
    if ext == ".csv":
        with open(abs_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
        return [h.strip() for h in headers if h]
    else:
        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            val = cell.value
            headers.append(str(val).strip() if val is not None else "")
        wb.close()
        return [h for h in headers if h]


def _iter_rows(file_path: str, skip_header: bool = True):
    abs_path = _resolve_path(file_path)
    ext = os.path.splitext(abs_path)[1].lower()
    if ext == ".csv":
        with open(abs_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            if skip_header:
                next(reader, None)
            for row_num, row in enumerate(reader, start=2):
                cells = [cell.strip() if cell else "" for cell in row]
                if not any(cells):
                    continue
                yield row_num, cells
    else:
        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=2 if skip_header else 1)
        for row_num, row in enumerate(rows, start=2 if skip_header else 1):
            cells = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
            if not any(cells):
                continue
            yield row_num, cells
        wb.close()


def _build_field_to_col(column_mapping: dict[str, str] | None) -> dict[str, str]:
    """Build reverse mapping: property_field -> source_column_name."""
    field_to_col: dict[str, str] = {}
    if column_mapping:
        for src_col, target_field in column_mapping.items():
            if target_field:
                field_to_col[target_field] = src_col
    return field_to_col


def _get_value(row_dict: dict[str, str | None], field: str, field_to_col: dict[str, str]) -> str | None:
    src_col = field_to_col.get(field)
    if src_col:
        return row_dict.get(src_col)
    return row_dict.get(field, row_dict.get(field.lower()))


def _validate_row(
    row_dict: dict[str, str | None],
    column_mapping: dict[str, str] | None,
    existing_codes: set[str],
) -> list[str]:
    errors: list[str] = []
    field_to_col = _build_field_to_col(column_mapping)

    def _get(field: str) -> str | None:
        if column_mapping:
            src_col = field_to_col.get(field)
            if src_col:
                return row_dict.get(src_col)
            return None
        return row_dict.get(field, row_dict.get(field.lower()))

    for req_field in REQUIRED_FIELDS:
        val = _get(req_field)
        if val is None or str(val).strip() == "":
            errors.append(f"import.missing_required||field={req_field}")

    prop_code = _get("property_code")
    if prop_code and str(prop_code).strip():
        if prop_code in existing_codes:
            errors.append(f"import.code_duplicate||code={prop_code}")

    rent_val = _get("monthly_rent")
    if rent_val and str(rent_val).strip():
        cleaned = _clean_price(rent_val)
        if cleaned is None:
            errors.append(f"import.invalid_rent||val={rent_val}")

    lat_val = _get("latitude")
    lng_val = _get("longitude")
    if lat_val and lng_val and str(lat_val).strip() and str(lng_val).strip():
        lat_f = _clean_float(lat_val)
        lng_f = _clean_float(lng_val)
        if lat_f is not None and lng_f is not None and not validate_latlng(lat_f, lng_f):
            errors.append(f"import.gps_out_of_range||lat={lat_f}||lng={lng_f}")

    for field in ["bedroom_count", "bathroom_count", "deposit_months"]:
        val = _get(field)
        if val and str(val).strip():
            if _clean_int(val) is None:
                errors.append(f"import.should_be_int||field={field}||val={val}")

    for field in ["size_sqm"]:
        val = _get(field)
        if val and str(val).strip():
            if _clean_float(val) is None:
                errors.append(f"import.should_be_number||field={field}||val={val}")

    return errors


# ─── public service functions ──────────────────────────────────────

async def upload_file(
    db: AsyncSession,
    file: UploadFile,
    current_user: User,
) -> ImportJob:
    original_name = file.filename or "unknown.csv"
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"import.unsupported_type||ext={ext}||allowed={', '.join(ALLOWED_EXTENSIONS)}",
        )

    import_dir = os.path.abspath(os.path.join(settings.UPLOAD_DIR, "..", "imports"))
    os.makedirs(import_dir, exist_ok=True)

    file_id = _uuid.uuid4()
    filename = f"{file_id}{ext}"
    filepath = os.path.join(import_dir, filename)

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"import.file_too_large||max_size={settings.MAX_UPLOAD_SIZE_MB}",
        )

    with open(filepath, "wb") as f:
        f.write(content)

    # Count rows
    total_rows = 0
    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            total_rows = sum(1 for _ in f) - 1
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        total_rows = max(ws.max_row - 1, 0)
        wb.close()

    job = ImportJob(
        original_filename=original_name,
        status="uploaded",
        created_by=current_user.id,
        total_rows=total_rows or 0,
        file_path=f"/static/imports/{filename}",
    )
    db.add(job)
    await db.flush()
    return job


async def get_columns(db: AsyncSession, import_job_id: UUID) -> dict:
    result = await db.execute(select(ImportJob).where(ImportJob.id == import_job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="import.job_not_found")

    abs_path = _resolve_path(job.file_path)
    headers = _read_headers(job.file_path)

    columns = []
    for h in headers:
        detected = _detect_column(h)
        columns.append({"header": h, "auto_detected_field": detected})

    sheet_names: list[str] = []
    ext = os.path.splitext(abs_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

    return {
        "columns": columns,
        "sheet_names": sheet_names,
        "total_rows": job.total_rows,
    }


async def preview(
    db: AsyncSession,
    import_job_id: UUID,
    mapping: dict[str, str] | None = None,
    limit: int = 50,
) -> dict:
    result = await db.execute(select(ImportJob).where(ImportJob.id == import_job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="import.job_not_found")

    headers = _read_headers(job.file_path)

    existing_result = await db.execute(select(Property.property_code).where(Property.is_deleted == False))
    existing_codes = set(existing_result.scalars().all())

    rows = []
    for row_num, cells in _iter_rows(job.file_path):
        if len(rows) >= limit:
            break
        row_dict: dict[str, str | None] = {}
        for i, header in enumerate(headers):
            val = cells[i] if i < len(cells) else None
            row_dict[header] = val

        errors = _validate_row(row_dict, mapping, existing_codes)
        rows.append({"row_number": row_num, "data": row_dict, "errors": errors})

    columns = []
    for h in headers:
        detected = _detect_column(h)
        columns.append({"header": h, "auto_detected_field": detected})

    return {
        "columns": columns,
        "rows": rows,
        "total_rows": job.total_rows,
        "preview_count": len(rows),
    }


async def map_fields(db: AsyncSession, import_job_id: UUID, mapping: dict[str, str]) -> ImportJob:
    result = await db.execute(select(ImportJob).where(ImportJob.id == import_job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="import.job_not_found")

    job.column_mapping = mapping
    job.status = "mapped"
    await db.flush()
    return job


async def confirm_import(
    db: AsyncSession,
    import_job_id: UUID,
    current_user: User,
    overwrite_existing: bool = False,
) -> dict:
    result = await db.execute(select(ImportJob).where(ImportJob.id == import_job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="import.job_not_found")

    headers = _read_headers(job.file_path)
    column_mapping: dict[str, str] = job.column_mapping or {}
    field_to_col = _build_field_to_col(column_mapping)

    def _get(row_dict: dict[str, str | None], field: str) -> str | None:
        return _get_value(row_dict, field, field_to_col)

    existing_result = await db.execute(select(Property.property_code).where(Property.is_deleted == False))
    existing_codes: set[str] = set(existing_result.scalars().all())

    update_map: dict[str, Property] = {}
    if overwrite_existing:
        props_result = await db.execute(select(Property).where(Property.property_code.in_(existing_codes)))
        update_map = {p.property_code: p for p in props_result.scalars().all()}

    job.status = "importing"
    await db.flush()

    total_rows = 0
    success_rows = 0
    error_rows = 0

    for row_num, cells in _iter_rows(job.file_path):
        total_rows += 1
        row_dict: dict[str, str | None] = {}
        for i, header in enumerate(headers):
            val = cells[i] if i < len(cells) else None
            row_dict[header] = val

        errors = _validate_row(row_dict, column_mapping, existing_codes)
        if errors:
            error_rows += 1
            err = ImportError(
                import_job_id=import_job_id,
                row_number=row_num,
                raw_data=row_dict,
                error_messages=errors,
            )
            db.add(err)
            continue

        try:
            name_val = str(_get(row_dict, "name") or "")
            address_val = str(_get(row_dict, "address") or "")
            status_val = str(_get(row_dict, "status") or "available")
            prop_code = str(_get(row_dict, "property_code") or "")
            if not prop_code:
                prop_code = f"IMPORT-{_uuid.uuid4().hex[:8].upper()}"

            rent_raw = _get(row_dict, "monthly_rent")
            rent_val = _clean_price(rent_raw) if rent_raw else None

            lat_raw = _get(row_dict, "latitude")
            lng_raw = _get(row_dict, "longitude")
            lat_val = _clean_float(lat_raw) if lat_raw else None
            lng_val = _clean_float(lng_raw) if lng_raw else None

            bedrooms_val = _clean_int(_get(row_dict, "bedroom_count")) if _get(row_dict, "bedroom_count") else None
            bathrooms_val = _clean_int(_get(row_dict, "bathroom_count")) if _get(row_dict, "bathroom_count") else None
            size_val = _clean_float(_get(row_dict, "size_sqm")) if _get(row_dict, "size_sqm") else None
            deposit_val = _clean_int(_get(row_dict, "deposit_months")) if _get(row_dict, "deposit_months") else None
            pet_val = _clean_bool(_get(row_dict, "pet_allowed"))
            tags_raw = _get(row_dict, "tags")
            tags_val = [t.strip() for t in str(tags_raw).split(",") if t.strip()] if tags_raw else None

            if overwrite_existing and prop_code in update_map:
                existing_prop = update_map[prop_code]
                existing_prop.name = name_val
                existing_prop.address = address_val
                existing_prop.status = status_val
                existing_prop.building_name = str(_get(row_dict, "building_name") or "") if _get(row_dict, "building_name") else None
                existing_prop.latitude = lat_val
                existing_prop.longitude = lng_val
                existing_prop.district = str(_get(row_dict, "district") or "") if _get(row_dict, "district") else None
                existing_prop.area = str(_get(row_dict, "area") or "") if _get(row_dict, "area") else None
                existing_prop.nearest_bts = str(_get(row_dict, "nearest_bts") or "") if _get(row_dict, "nearest_bts") else None
                existing_prop.nearest_mrt = str(_get(row_dict, "nearest_mrt") or "") if _get(row_dict, "nearest_mrt") else None
                existing_prop.bedroom_count = bedrooms_val
                existing_prop.bathroom_count = bathrooms_val
                existing_prop.size_sqm = size_val
                existing_prop.monthly_rent = rent_val
                existing_prop.deposit_months = deposit_val
                existing_prop.pet_allowed = pet_val or False
                existing_prop.contact_person = str(_get(row_dict, "contact_person") or "") if _get(row_dict, "contact_person") else None
                existing_prop.contact_phone = str(_get(row_dict, "contact_phone") or "") if _get(row_dict, "contact_phone") else None
                existing_prop.contact_line = str(_get(row_dict, "contact_line") or "") if _get(row_dict, "contact_line") else None
                existing_prop.description = str(_get(row_dict, "description") or "") if _get(row_dict, "description") else None
                existing_prop.internal_note = str(_get(row_dict, "internal_note") or "") if _get(row_dict, "internal_note") else None
                existing_prop.tags = tags_val
            else:
                prop = Property(
                    property_code=prop_code,
                    name=name_val,
                    building_name=str(_get(row_dict, "building_name") or "") if _get(row_dict, "building_name") else None,
                    address=address_val,
                    latitude=lat_val,
                    longitude=lng_val,
                    district=str(_get(row_dict, "district") or "") if _get(row_dict, "district") else None,
                    area=str(_get(row_dict, "area") or "") if _get(row_dict, "area") else None,
                    nearest_bts=str(_get(row_dict, "nearest_bts") or "") if _get(row_dict, "nearest_bts") else None,
                    nearest_mrt=str(_get(row_dict, "nearest_mrt") or "") if _get(row_dict, "nearest_mrt") else None,
                    bedroom_count=bedrooms_val,
                    bathroom_count=bathrooms_val,
                    size_sqm=size_val,
                    monthly_rent=rent_val,
                    deposit_months=deposit_val,
                    status=status_val,
                    pet_allowed=pet_val or False,
                    contact_person=str(_get(row_dict, "contact_person") or "") if _get(row_dict, "contact_person") else None,
                    contact_phone=str(_get(row_dict, "contact_phone") or "") if _get(row_dict, "contact_phone") else None,
                    contact_line=str(_get(row_dict, "contact_line") or "") if _get(row_dict, "contact_line") else None,
                    description=str(_get(row_dict, "description") or "") if _get(row_dict, "description") else None,
                    internal_note=str(_get(row_dict, "internal_note") or "") if _get(row_dict, "internal_note") else None,
                    tags=tags_val,
                    created_by=current_user.id,
                )
                db.add(prop)
                existing_codes.add(prop_code)

            success_rows += 1

        except Exception as e:
            error_rows += 1
            err = ImportError(
                import_job_id=import_job_id,
                row_number=row_num,
                raw_data=row_dict,
                error_messages=[f"导入异常: {str(e)}"],
            )
            db.add(err)

    job.total_rows = total_rows
    job.success_rows = success_rows
    job.error_rows = error_rows
    job.status = "imported"
    await db.flush()

    # n8n: notify import complete
    from app.services.n8n_service import trigger_import_complete
    await trigger_import_complete(
        db,
        str(import_job_id),
        job.original_filename,
        total_rows,
        success_rows,
        error_rows,
    )

    # Trigger geocoding backfill for newly imported properties without GPS
    await _backfill_geocoding(db)

    return {
        "import_job_id": import_job_id,
        "total_rows": total_rows,
        "success_rows": success_rows,
        "error_rows": error_rows,
        "status": "imported",
    }


async def get_import_jobs(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[ImportJob], int]:
    query = select(ImportJob)
    count_query = select(func.count(ImportJob.id))

    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    query = query.order_by(ImportJob.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = list(result.scalars().all())
    return items, total


async def get_import_job(db: AsyncSession, import_job_id: UUID) -> ImportJob:
    result = await db.execute(select(ImportJob).where(ImportJob.id == import_job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="import.job_not_found")
    return job


async def get_import_errors(db: AsyncSession, import_job_id: UUID) -> list[ImportError]:
    result = await db.execute(
        select(ImportError)
        .where(ImportError.import_job_id == import_job_id)
        .order_by(ImportError.row_number)
    )
    return list(result.scalars().all())


async def _backfill_geocoding(db: AsyncSession) -> None:
    """After import, geocode up to 20 new properties without GPS coordinates."""
    from app.utils.geo import geocode_google, geocode_nominatim, validate_latlng

    result = await db.execute(
        select(Property).where(
            Property.is_deleted == False,
            Property.address.isnot(None),
            Property.address != "",
            Property.latitude.is_(None),
        ).limit(20)
    )
    props = list(result.scalars().all())

    for prop in props:
        if not prop.address:
            continue
        try:
            results = await geocode_google(prop.address)
            if not results:
                results = await geocode_nominatim(prop.address)
            if results:
                best = results[0]
                if validate_latlng(best.latitude, best.longitude):
                    prop.latitude = best.latitude
                    prop.longitude = best.longitude
        except Exception:
            pass

    await db.commit()
